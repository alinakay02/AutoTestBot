import os
import re
import json
import time
import shutil
import zipfile
import logging
from dataclasses import dataclass
from typing import Callable, Optional, List, Tuple

from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, WebDriverException


# кнопка обновления (запускает скрипт обновления)
X_REFRESH_BTN = (
    "/html/body/div[1]/div[1]/div[2]/div[3]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr/td/"
    "table/tbody/tr/td/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div[1]/div/div/div[2]/a"
)

# paging label
X_PAGING_INFO = (
    "/html/body/div[1]/div[1]/div[2]/div[3]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr/td/"
    "table/tbody/tr/td/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div[1]/div/div/div[3]/"
    "table/tbody/tr/td/table/tbody/tr/td/span"
)

# next page button
X_NEXT_PAGE_BTN = (
    "/html/body/div[1]/div[1]/div[2]/div[3]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr/td/"
    "table/tbody/tr/td/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[3]/div[1]/div/div/div/div/"
    "div[6]/div/button"
)

# tbody with rows
X_TABLE_TBODY = (
    "/html/body/div[1]/div[1]/div[2]/div[3]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr/td/"
    "table/tbody/tr/td/div[1]/div[1]/div/div/div/div[1]/div/div/div/div[2]/table/tbody[1]"
)

# row selection cell and guid cell (relative)
REL_TD_SELECT = "./td[1]"
REL_TD_GUID = "./td[9]"

# export button (same as in table_export2)
X_BTN_EXPORT_TXT = (
    "/html/body/div[1]/div[1]/div[2]/div[3]/div[2]/div/div/div/div/div[2]/div[1]/div[1]/div/div/div[2]/"
    "div/div[3]/div/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td[9]/button"
)


@dataclass
class WaitCfg:
    short: int = 3
    medium: int = 9
    long: int = 15
    poll: float = 0.2


def _now() -> float:
    return time.time()


def _safe_sleep(seconds: float, stop_check: Optional[Callable[[], bool]] = None):
    t0 = _now()
    while _now() - t0 < seconds:
        if stop_check and stop_check():
            return
        time.sleep(0.1)


def _wait(driver, timeout: int, poll: float):
    return WebDriverWait(driver, timeout, poll_frequency=poll)


def _find(driver, by, selector, timeout: int, poll: float):
    """Ищем в текущем контексте (таблица может быть во фрейме)."""
    return _wait(driver, timeout, poll).until(EC.presence_of_element_located((by, selector)))


def _find_clickable(driver, by, selector, timeout: int, poll: float):
    """Ищем в текущем контексте."""
    return _wait(driver, timeout, poll).until(EC.element_to_be_clickable((by, selector)))


def _scroll_into_view(driver, el):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
    except Exception:
        pass


def _robust_click(driver, el) -> bool:
    try:
        _scroll_into_view(driver, el)
        time.sleep(0.1)
    except Exception:
        pass

    try:
        el.click()
        return True
    except WebDriverException:
        pass

    try:
        ActionChains(driver).move_to_element(el).pause(0.05).click().perform()
        return True
    except WebDriverException:
        pass

    try:
        driver.execute_script("arguments[0].click();", el)
        return True
    except WebDriverException:
        return False


def _list_files(dir_path: str) -> List[str]:
    try:
        return [os.path.join(dir_path, f) for f in os.listdir(dir_path)]
    except Exception:
        return []


def _is_partial_download(path: str) -> bool:
    low = path.lower()
    return low.endswith(".crdownload") or low.endswith(".tmp") or low.endswith(".part")


def _wait_for_new_zip(download_dir: str, since_ts: float, timeout: int, stop_check=None) -> Optional[str]:
    deadline = _now() + timeout
    last_candidate = None
    last_size = None
    stable_since = None

    while _now() < deadline:
        if stop_check and stop_check():
            return None

        candidates = []
        for p in _list_files(download_dir):
            try:
                if os.path.isdir(p) or _is_partial_download(p):
                    continue
                if not p.lower().endswith(".zip"):
                    continue
                if os.path.getmtime(p) >= since_ts - 0.2:
                    candidates.append(p)
            except Exception:
                continue

        if candidates:
            candidates.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            cand = candidates[0]

            try:
                size = os.path.getsize(cand)
            except Exception:
                size = None

            if cand != last_candidate:
                last_candidate = cand
                last_size = size
                stable_since = _now()
            else:
                if size is not None and last_size is not None and size == last_size:
                    if stable_since and (_now() - stable_since) >= 1.2:
                        return cand
                else:
                    last_size = size
                    stable_since = _now()

        time.sleep(0.25)

    return None


def _unique_txt_path(dst_dir: str, base_name: str) -> str:
    base = base_name.strip()
    if not base:
        base = "unknown_guid"
    filename = f"{base}.txt"
    candidate = os.path.join(dst_dir, filename)
    if not os.path.exists(candidate):
        return candidate
    i = 1
    while True:
        candidate = os.path.join(dst_dir, f"{base}__{i}.txt")
        if not os.path.exists(candidate):
            return candidate
        i += 1


def _extract_first_txt(zip_path: str, dst_dir: str) -> Optional[str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = [n for n in zf.namelist() if not n.endswith("/") and n.lower().endswith(".txt")]
            if not names:
                return None
            name = names[0]
            zf.extract(name, dst_dir)
            return os.path.join(dst_dir, name)
    except Exception:
        return None


def _paging_parse(text: str) -> Tuple[int, int, int, int]:
    """
    Пример: 'Отображено: 1 из 1 страниц (6 из 6 записей)'
    Возвращает: current_page, total_pages, shown_count, total_records
    """
    # current page / total pages
    m_pages = re.search(r"Отображено:\s*(\d+)\s*из\s*(\d+)\s*страниц", text)
    # shown / total records
    m_rec = re.search(r"\(\s*(\d+)\s*из\s*(\d+)\s*запис", text)

    cur = int(m_pages.group(1)) if m_pages else 1
    tot = int(m_pages.group(2)) if m_pages else 1
    shown = int(m_rec.group(1)) if m_rec else 0
    total = int(m_rec.group(2)) if m_rec else 0
    return cur, tot, shown, total


def _click_refresh_and_wait(driver, cfg: WaitCfg, stop_check=None) -> bool:
    """Нажимает кнопку обновления, ждёт 2–3 сек."""
    try:
        btn = _find_clickable(driver, By.XPATH, X_REFRESH_BTN, cfg.medium, cfg.poll)
        if _robust_click(driver, btn):
            _safe_sleep(2.5, stop_check)
            return True
    except (TimeoutException, Exception):
        pass
    return False


def _paging_has_error(txt: str) -> bool:
    """Ошибка: вместо цифр отображается '?'."""
    return "?" in (txt or "")


def get_paging_info(driver, cfg: WaitCfg) -> Tuple[int, int, int, int, str]:
    """Считывает label пагинации и парсит числа"""
    el = _find(driver, By.XPATH, X_PAGING_INFO, cfg.medium, cfg.poll)
    txt = (el.text or "").strip()
    cur, tot, shown, total = _paging_parse(txt)
    return cur, tot, shown, total, txt


def get_paging_info_with_retry(driver, cfg: WaitCfg, stop_check=None) -> Tuple[int, int, int, int, str]:
    """
    Клик обновления, ожидание, получение пагинации.
    При ошибке ('?' в тексте) — повторить клик/ожидание/проверку до 3 раз.
    """
    _click_refresh_and_wait(driver, cfg, stop_check)

    for attempt in range(3):
        if stop_check and stop_check():
            raise RuntimeError("Остановлено пользователем")

        cur, tot, shown, total, txt = get_paging_info(driver, cfg)

        if not _paging_has_error(txt):
            return cur, tot, shown, total, txt

        if attempt < 2:
            _click_refresh_and_wait(driver, cfg, stop_check)
            _safe_sleep(0.5, stop_check)

    return cur, tot, shown, total, txt


def _get_rows(driver, cfg: WaitCfg):
    tbody = _find(driver, By.XPATH, X_TABLE_TBODY, cfg.medium, cfg.poll)
    return tbody.find_elements(By.XPATH, "./tr")


def _row_is_selected(tr_el) -> bool:
    try:
        cls = (tr_el.get_attribute("class") or "").lower()
        return "z-listitem-selected" in cls
    except Exception:
        return False


def _click_row_select_cell(driver, tr_el) -> bool:
    try:
        td = tr_el.find_element(By.XPATH, REL_TD_SELECT)
    except Exception:
        return False
    return _robust_click(driver, td)


def _ensure_row_selected(driver, tr_el, cfg: WaitCfg, stop_check=None) -> bool:
    """Выделяет строку (до 3 попыток клика по ячейке и ожидания класса z-listitem-selected)."""
    if _row_is_selected(tr_el):
        return True
    for attempt in range(1, 4):
        if stop_check and stop_check():
            return False
        if not _click_row_select_cell(driver, tr_el):
            if attempt < 3:
                _safe_sleep(0.5, stop_check)
            continue
        deadline = _now() + cfg.short
        while _now() < deadline:
            if stop_check and stop_check():
                return False
            try:
                if _row_is_selected(tr_el):
                    return True
            except StaleElementReferenceException:
                return True
            time.sleep(0.15)
        try:
            if _row_is_selected(tr_el):
                return True
        except Exception:
            pass
        if attempt < 3:
            _safe_sleep(0.5, stop_check)
    try:
        return _row_is_selected(tr_el)
    except Exception:
        return False


def _ensure_row_unselected(driver, tr_el, cfg: WaitCfg, stop_check=None) -> bool:
    if not _row_is_selected(tr_el):
        return True
    if not _click_row_select_cell(driver, tr_el):
        return False
    deadline = _now() + cfg.short
    while _now() < deadline:
        if stop_check and stop_check():
            return False
        try:
            if not _row_is_selected(tr_el):
                return True
        except StaleElementReferenceException:
            return True
        time.sleep(0.15)
    try:
        return not _row_is_selected(tr_el)
    except Exception:
        return True


def _read_guid_from_row(tr_el) -> str:
    try:
        td = tr_el.find_element(By.XPATH, REL_TD_GUID)
        title = (td.get_attribute("title") or "").strip()
        if title:
            return title
        return (td.text or "").strip()
    except Exception:
        return ""


def _go_next_page(driver, cfg: WaitCfg, stop_check=None) -> bool:
    try:
        before = get_paging_info(driver, cfg)[4]
    except Exception:
        before = ""

    try:
        btn = _find_clickable(driver, By.XPATH, X_NEXT_PAGE_BTN, cfg.medium, cfg.poll)
    except TimeoutException:
        return False

    if not _robust_click(driver, btn):
        return False

    deadline = _now() + cfg.medium
    while _now() < deadline:
        if stop_check and stop_check():
            return False
        try:
            after = get_paging_info(driver, cfg)[4]
            if after and after != before:
                return True
        except Exception:
            pass
        time.sleep(0.2)

    return True


def _progress_path(txt_out_dir: str) -> str:
    return os.path.join(txt_out_dir, "_progress.json")


GUIDS_EXCEL_FILENAME = "processed_guids.xlsx"


def _guids_excel_path(txt_out_dir: str) -> str:
    return os.path.join(txt_out_dir, GUIDS_EXCEL_FILENAME)


def _init_guids_excel_for_export(txt_out_dir: str, start_index: int) -> None:
    """
    Перед экспортом: если начинаем с 1 — удаляем старый файл GUID
    Если с N > 1 — файл не трогаем, запись продолжится с нужной строки
    """
    if start_index <= 1:
        p = _guids_excel_path(txt_out_dir)
        try:
            if os.path.isfile(p):
                os.remove(p)
        except Exception:
            pass


def _write_guid_to_excel(txt_out_dir: str, record_index_1based: int, guid: str) -> None:
    """Записывает GUID обработанной записи в строку record_index_1based (1-based) в Excel. Строка 1 — заголовок"""
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        logging.warning("openpyxl не установлен: GUID не записываются в Excel")
        return
    p = _guids_excel_path(txt_out_dir)
    excel_row = record_index_1based + 1  # строка 1 — заголовок "GUID"
    try:
        if os.path.isfile(p):
            wb = load_workbook(p)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "GUID"
            ws.cell(row=1, column=1, value="GUID")
        ws.cell(row=excel_row, column=1, value=guid)
        wb.save(p)
    except Exception as e:
        logging.warning("Не удалось записать GUID в Excel: %s", e)


def load_progress(txt_out_dir: str) -> int:
    """Возвращает номер последней успешно обработанной записи (1-based), либо 0"""
    p = _progress_path(txt_out_dir)
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        return int(data.get("last_done", 0))
    except Exception:
        return 0


def save_progress(txt_out_dir: str, last_done: int):
    """Сохраняет номер последней успешно обработанной записи (1-based)"""
    p = _progress_path(txt_out_dir)
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump({"last_done": int(last_done)}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def ask_start_index(default_start: int) -> int:
    """Спрашивает с какой записи начинать"""
    try:
        s = input(f"С какой записи начинать скачивание (1 — с начала)? [по умолчанию {default_start}]: ").strip()
    except EOFError:
        s = ""
    if not s:
        start = default_start
    else:
        try:
            start = int(s)
        except Exception:
            start = default_start
    if start <= 0:
        start = 1
    return start


def export_all_rows_to_txt(
    driver,
    download_dir: str,
    txt_out_dir: str,
    cfg: WaitCfg = WaitCfg(),
    stop_check=None,
    start_index: int = 1,
) -> Tuple[int, int]:
    """
    Выгружает TXT для каждой строки всех страниц
    start_index задаётся снаружи (запрос в начале программы). 0 недопустим — передавать не меньше 1
    Возвращает (total_records, downloaded_count)
    """
    if start_index <= 0:
        start_index = 1
    original_implicit = None
    total_records_stored = 0  # общее число записей для финального вывода
    downloaded = 0
    try:
        try:
            original_implicit = driver.timeouts.implicit_wait
        except Exception:
            original_implicit = None
        try:
            driver.implicitly_wait(0)
        except Exception:
            pass

        os.makedirs(txt_out_dir, exist_ok=True)
        _init_guids_excel_for_export(txt_out_dir, start_index)

        cur_page, total_pages, shown, total_records, _ = get_paging_info_with_retry(
            driver, cfg, stop_check
        )

        if total_records <= 0:
            # fallback: считаем по страницам/строкам первой страницы
            rows_first = len(_get_rows(driver, cfg))
            total_records = total_pages * rows_first

        total_records_stored = total_records  # для финального вывода "Всего M записей"

        skipped = max(0, start_index - 1)
        total_effective = max(0, total_records - skipped)

        # вычисляем на какую страницу перейти и с какой строки на странице начать
        rows_on_first_page = max(1, len(_get_rows(driver, cfg)))
        page_size = rows_on_first_page
        target_page = ((start_index - 1) // page_size) + 1
        start_row_in_page = ((start_index - 1) % page_size) + 1

        # дойти до target_page
        while cur_page < target_page:
            if stop_check and stop_check():
                return total_records_stored, 0
            if not _go_next_page(driver, cfg, stop_check):
                raise RuntimeError(f"Не удалось перейти на страницу {cur_page + 1}")
            cur_page, _, _, _, _ = get_paging_info(driver, cfg)
            _safe_sleep(0.6, stop_check)

        global_index = (cur_page - 1) * page_size + 1

        # на целевой странице корректируем глобальный индекс
        if cur_page == target_page:
            global_index = start_index

        while True:
            if stop_check and stop_check():
                break

            rows = _get_rows(driver, cfg)
            if not rows:
                break

            # вычисляем с какого индекса в rows стартовать
            row_start = 1
            if cur_page == target_page:
                row_start = start_row_in_page
            if global_index > start_index:
                row_start = 1

            for r_idx in range(row_start, len(rows) + 1):
                if stop_check and stop_check():
                    break

                # освежаем список строк чтобы не ловить stale
                rows = _get_rows(driver, cfg)
                if r_idx - 1 >= len(rows):
                    break
                tr = rows[r_idx - 1]

                # три попытки на одну запись
                ok_one = False
                last_error = None
                for attempt in range(1, 4):
                    if stop_check and stop_check():
                        break
                    try:
                        if not _ensure_row_selected(driver, tr, cfg, stop_check):
                            raise RuntimeError("Не удалось выделить строку")

                        guid = _read_guid_from_row(tr)
                        if not guid:
                            raise RuntimeError("GUID пустой или не найден")

                        since_ts = _now()

                        btn = _find_clickable(driver, By.XPATH, X_BTN_EXPORT_TXT, cfg.medium, cfg.poll)
                        if not _robust_click(driver, btn):
                            raise RuntimeError("Не удалось нажать экспорт TXT")

                        zip_path = _wait_for_new_zip(download_dir, since_ts, cfg.long, stop_check)
                        if not zip_path or not os.path.exists(zip_path):
                            raise RuntimeError("Не удалось дождаться нового ZIP")

                        extracted_path = _extract_first_txt(zip_path, txt_out_dir)
                        if not extracted_path or not os.path.exists(extracted_path):
                            raise RuntimeError("TXT не найден в ZIP или не извлечён")

                        dst_txt = _unique_txt_path(txt_out_dir, guid)
                        shutil.move(extracted_path, dst_txt)

                        save_progress(txt_out_dir, global_index)
                        _write_guid_to_excel(txt_out_dir, global_index, guid)

                        if not _ensure_row_unselected(driver, tr, cfg, stop_check):
                            raise RuntimeError("Не удалось снять выделение строки")

                        downloaded += 1
                        ok_one = True
                        break

                    except Exception as e:
                        last_error = e
                        try:
                            _ensure_row_unselected(driver, tr, cfg, stop_check)
                        except Exception:
                            pass
                        _safe_sleep(0.8, stop_check)

                if not ok_one:
                    # Восстановление при сбое (например, в систему пришла новая запись): обновить, перейти на нужную страницу, повторить 3 попытки
                    _click_refresh_and_wait(driver, cfg, stop_check)
                    _safe_sleep(0.6, stop_check)
                    cur_page, tot_pages_now, shown_now, total_now, _ = get_paging_info(driver, cfg)
                    total_pages = tot_pages_now or total_pages
                    target_page = ((global_index - 1) // page_size) + 1
                    while cur_page < target_page:
                        if stop_check and stop_check():
                            break
                        if not _go_next_page(driver, cfg, stop_check):
                            break
                        cur_page, _, _, _, _ = get_paging_info(driver, cfg)
                        _safe_sleep(0.8, stop_check)
                        _click_refresh_and_wait(driver, cfg, stop_check)
                    rows = _get_rows(driver, cfg)
                    row_in_page = ((global_index - 1) % page_size) + 1
                    if row_in_page > len(rows):
                        err_text = str(last_error or "")
                        if "выделить строку" in err_text.lower():
                            logging.warning(
                                "Запись #%s: не удалось выделить строку после 3 попыток (в каждой попытке — до 3 кликов)",
                                global_index,
                            )
                        print(f"Всего {total_records_stored} записей. Скачано {downloaded} записей.")
                        print(f"ОШИБКА. Последняя успешно обработанная запись: {load_progress(txt_out_dir)}")
                        raise RuntimeError(f"Не удалось обработать запись #{global_index}: {last_error}")
                    tr = rows[row_in_page - 1]
                    last_error = None
                    for attempt in range(1, 4):
                        if stop_check and stop_check():
                            break
                        try:
                            if not _ensure_row_selected(driver, tr, cfg, stop_check):
                                raise RuntimeError("Не удалось выделить строку")
                            guid = _read_guid_from_row(tr)
                            if not guid:
                                raise RuntimeError("GUID пустой или не найден")
                            since_ts = _now()
                            btn = _find_clickable(driver, By.XPATH, X_BTN_EXPORT_TXT, cfg.medium, cfg.poll)
                            if not _robust_click(driver, btn):
                                raise RuntimeError("Не удалось нажать экспорт TXT")
                            zip_path = _wait_for_new_zip(download_dir, since_ts, cfg.long, stop_check)
                            if not zip_path or not os.path.exists(zip_path):
                                raise RuntimeError("Не удалось дождаться нового ZIP")
                            extracted_path = _extract_first_txt(zip_path, txt_out_dir)
                            if not extracted_path or not os.path.exists(extracted_path):
                                raise RuntimeError("TXT не найден в ZIP или не извлечён")
                            dst_txt = _unique_txt_path(txt_out_dir, guid)
                            shutil.move(extracted_path, dst_txt)
                            save_progress(txt_out_dir, global_index)
                            _write_guid_to_excel(txt_out_dir, global_index, guid)
                            if not _ensure_row_unselected(driver, tr, cfg, stop_check):
                                raise RuntimeError("Не удалось снять выделение строки")
                            downloaded += 1
                            ok_one = True
                            break
                        except Exception as e:
                            last_error = e
                            try:
                                _ensure_row_unselected(driver, tr, cfg, stop_check)
                            except Exception:
                                pass
                            _safe_sleep(0.8, stop_check)
                    if not ok_one:
                        err_text = str(last_error or "")
                        if "выделить строку" in err_text.lower():
                            logging.warning(
                                "Запись #%s: не удалось выделить строку после 3 попыток (в каждой попытке — до 3 кликов)",
                                global_index,
                            )
                        print(f"Всего {total_records_stored} записей. Скачано {downloaded} записей.")
                        print(f"ОШИБКА. Последняя успешно обработанная запись: {load_progress(txt_out_dir)}")
                        raise RuntimeError(f"Не удалось обработать запись #{global_index}: {last_error}")

                global_index += 1

                # Последняя строка на странице: ждём полного завершения перед переходом на следующую страницу
                if r_idx == len(rows):
                    _safe_sleep(1.5, stop_check)

            if stop_check and stop_check():
                break

            # если страниц больше нет
            cur_page, tot_pages_now, shown_now, total_now, _ = get_paging_info(driver, cfg)
            total_pages = tot_pages_now or total_pages
            total_records = total_now or total_records
            # на последней странице уточняем общее число записей
            if tot_pages_now > 1 and cur_page >= tot_pages_now and shown_now > 0:
                total_records_stored = (tot_pages_now - 1) * rows_on_first_page + shown_now

            if cur_page >= total_pages:
                break

            if not _go_next_page(driver, cfg, stop_check):
                break

            cur_page, _, _, _, _ = get_paging_info(driver, cfg)
            _safe_sleep(0.8, stop_check)
            # После перехода на новую страницу нажимаем кнопку обновления
            _click_refresh_and_wait(driver, cfg, stop_check)

        print(f"Всего {total_records_stored} записей. Скачано {downloaded} записей.")
        return total_records_stored, downloaded

    finally:
        try:
            if original_implicit is not None:
                driver.implicitly_wait(int(original_implicit))
            else:
                driver.implicitly_wait(5)
        except Exception:
            pass
